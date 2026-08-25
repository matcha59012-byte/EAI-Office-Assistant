"""模块C：客户数据管理（复用信息提取的实体层）。

- 实体列表/搜索/筛选/详情（客户为主，含项目/公司）
- 编辑卡片（字段 + status/last_contact）
- 跟进记录（一键更新 last_contact=今天, status=跟进中）
- Excel 导入（name 必填，重名跳过）
- 静默告警（>30 天未跟进 且 未关闭）
- 看板统计
"""
from datetime import date, datetime
import json

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel
from sqlalchemy import or_
from sqlalchemy.orm import Session

from app.ai.cards import render_entity_card
from app.ai.llm import llm
from app.config import settings
from app.database import get_db
from app.models import Entity
from app.rag.embedder import embed_texts

router = APIRouter(prefix="/api/customer", tags=["customer"])

SILENT_DAYS = 30
STATUSES = ["新", "跟进中", "静默", "已关闭"]


class UpdateRequest(BaseModel):
    name: str | None = None
    card: dict | None = None
    card_md: str | None = None  # Obsidian 式 md 编辑保存
    status: str | None = None
    last_contact: str | None = None


class FollowRequest(BaseModel):
    status: str | None = "跟进中"


def _entity_out(e: Entity) -> dict:
    try:
        card = json.loads(e.card_json or "{}")
    except json.JSONDecodeError:
        card = {}
    return {
        "id": e.id,
        "entity_type": e.entity_type,
        "name": e.name,
        "card": card,
        "card_md": e.card_md,
        "status": e.status,
        "last_contact": e.last_contact,
        "source_meeting_id": e.source_meeting_id,
        "created_at": e.created_at,
    }


def _is_silent(e: Entity) -> bool:
    if e.status == "已关闭":
        return False
    if not e.last_contact:
        return True
    try:
        last = datetime.strptime(e.last_contact, "%Y-%m-%d").date()
    except ValueError:
        return True
    return (date.today() - last).days > SILENT_DAYS


@router.get("/entities")
def list_entities(
    entity_type: str = "customer",
    status: str | None = None,
    q: str | None = None,
    db: Session = Depends(get_db),
):
    query = db.query(Entity).filter(Entity.entity_type == entity_type)
    if status:
        query = query.filter(Entity.status == status)
    if q:
        like = f"%{q}%"
        query = query.filter(or_(Entity.name.like(like), Entity.card_json.like(like)))
    rows = query.order_by(Entity.created_at.desc()).all()
    return [_entity_out(e) for e in rows]


@router.get("/entities/{entity_id}")
def get_entity(entity_id: int, db: Session = Depends(get_db)):
    e = db.get(Entity, entity_id)
    if not e:
        raise HTTPException(status_code=404, detail="实体不存在")
    return _entity_out(e)


@router.put("/entities/{entity_id}")
def update_entity(entity_id: int, req: UpdateRequest, db: Session = Depends(get_db)):
    e = db.get(Entity, entity_id)
    if not e:
        raise HTTPException(status_code=404, detail="实体不存在")
    if req.card is not None:
        e.card_json = json.dumps(req.card, ensure_ascii=False)
        e.card_md = render_entity_card(e.entity_type, req.card, e.source_meeting_id)
    if req.card_md is not None:
        # Obsidian 式 md 编辑：直接保存 md 卡片正文
        e.card_md = req.card_md
    if req.name:
        e.name = req.name
    if req.status and req.status in STATUSES:
        e.status = req.status
    if req.last_contact is not None:
        e.last_contact = req.last_contact
    db.commit()
    return _entity_out(e)


@router.delete("/entities/{entity_id}")
def delete_entity(entity_id: int, db: Session = Depends(get_db)):
    e = db.get(Entity, entity_id)
    if not e:
        raise HTTPException(status_code=404, detail="实体不存在")
    db.delete(e)
    db.commit()
    return {"ok": True}


@router.post("/entities/{entity_id}/follow")
def follow_entity(entity_id: int, req: FollowRequest, db: Session = Depends(get_db)):
    e = db.get(Entity, entity_id)
    if not e:
        raise HTTPException(status_code=404, detail="实体不存在")
    e.last_contact = date.today().isoformat()
    if req.status and req.status in STATUSES:
        e.status = req.status
    elif e.status in ("新", "静默"):
        e.status = "跟进中"
    db.commit()
    return _entity_out(e)


@router.post("/import")
def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Excel 导入客户：必填 name，重名跳过。"""
    from io import BytesIO

    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Excel 解析失败: {exc}")
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Excel 为空")

    # 表头映射（注意：列号 0 在 Python 里是 falsy，不能用 or 链）
    headers = [str(c.value or "").strip() for c in ws[1]]
    col = {h: i for i, h in enumerate(headers) if h}

    def _c(cn: str, en: str) -> int | None:
        if cn in col:
            return col[cn]
        return col.get(en)

    name_col = _c("姓名", "name")
    if name_col is None:
        raise HTTPException(status_code=400, detail="Excel 需包含『姓名』列")
    field_map = {
        "name": name_col,
        "company": _c("公司", "company"),
        "phone": _c("电话", "phone"),
        "email": _c("邮箱", "email"),
        "role": _c("角色", "role"),
        "position": _c("岗位", "position"),
        "notes": _c("备注", "notes"),
    }

    imported = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[name_col]:
            continue
        name = str(row[name_col]).strip()
        exists = (
            db.query(Entity)
            .filter(Entity.entity_type == "customer", Entity.name == name)
            .first()
        )
        if exists:
            skipped += 1
            continue
        card = {}
        for key, idx in field_map.items():
            if idx is not None and idx < len(row) and row[idx] not in (None, ""):
                card[key] = str(row[idx]).strip()
        if not card.get("name"):
            continue
        md = render_entity_card("customer", card, None)
        db.add(
            Entity(
                entity_type="customer",
                name=name,
                card_json=json.dumps(card, ensure_ascii=False),
                card_md=md,
                status="新",
            )
        )
        imported += 1
    db.commit()
    return {"imported": imported, "skipped": skipped}


@router.get("/alerts")
def silent_alerts(db: Session = Depends(get_db)):
    """静默告警：>30天未跟进 且 未关闭。"""
    rows = db.query(Entity).filter(Entity.entity_type == "customer").all()
    alerts = []
    for e in rows:
        if _is_silent(e):
            days = None
            if e.last_contact:
                try:
                    days = (date.today() - datetime.strptime(e.last_contact, "%Y-%m-%d").date()).days
                except ValueError:
                    days = None
            try:
                company = json.loads(e.card_json or "{}").get("company") or ""
            except json.JSONDecodeError:
                company = ""
            alerts.append(
                {
                    "id": e.id,
                    "name": e.name,
                    "company": company,
                    "status": e.status,
                    "last_contact": e.last_contact,
                    "days": days,
                }
            )
    return alerts


@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db)):
    """看板：客户/公司/项目数量 + 状态分布 + 静默数。"""
    counts = {
        t: db.query(Entity).filter(Entity.entity_type == t).count()
        for t in ("customer", "project", "company")
    }
    customers = db.query(Entity).filter(Entity.entity_type == "customer").all()
    status_dist = {s: 0 for s in STATUSES}
    silent = 0
    for e in customers:
        status_dist[e.status] = status_dist.get(e.status, 0) + 1
        if _is_silent(e):
            silent += 1
    return {
        "customers": counts["customer"],
        "projects": counts["project"],
        "companies": counts["company"],
        "total": counts["customer"],
        "status_dist": status_dist,
        "silent": silent,
    }


class AskRequest(BaseModel):
    question: str


def _linked_names(card: dict) -> list[str]:
    names = []
    for key in ("linked_projects", "linked_companies", "linked_people"):
        v = card.get(key) or []
        if isinstance(v, list):
            names.extend(str(x) for x in v if x)
    return names


@router.post("/ask")
def ask_entity(req: AskRequest, db: Session = Depends(get_db)):
    """实体问答：客户数据=可问答的实体知识库。

    检索实体卡片（numpy 余弦）+ 关联展开 → DeepSeek 总结，带来源。
    """
    question = req.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="问题不能为空")
    if not settings.llm_configured:
        raise HTTPException(status_code=502, detail="未配置 DeepSeek API Key")

    rows = db.query(Entity).all()
    items = []
    for e in rows:
        md = e.card_md or ""
        if not md.strip():
            continue
        items.append({"id": e.id, "type": e.entity_type, "name": e.name, "md": md})

    if not items:
        return {"answer": "暂无实体数据，请先在信息提取模块提取实体。", "sources": []}

    try:
        import numpy as np

        vecs = embed_texts([i["md"] for i in items])
        M = np.asarray(vecs, dtype=np.float32)
        q = np.asarray(embed_texts([question])[0], dtype=np.float32)
        sims = (M @ q) / (np.linalg.norm(M, axis=1) * np.linalg.norm(q) + 1e-9)
        top_idx = np.argsort(-sims)[:4].tolist()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"检索失败: {exc}")

    top = [items[i] for i in top_idx]

    # 关联展开：把 top 卡片关联到的实体也纳入上下文
    context_items = list(top)
    linked_names = set()
    for it in top:
        try:
            card = json.loads(db.get(Entity, it["id"]).card_json or "{}")
        except Exception:  # noqa: BLE001
            card = {}
        linked_names.update(_linked_names(card))
    if linked_names:
        for e in db.query(Entity).all():
            if e.name in linked_names and e.id not in [x["id"] for x in context_items]:
                context_items.append({"id": e.id, "type": e.entity_type, "name": e.name, "md": e.card_md or ""})
            if len(context_items) >= 8:
                break

    context = "\n\n".join(
        f"【{'客户' if i['type']=='customer' else '项目' if i['type']=='project' else '公司'}】{i['name']}\n{i['md']}"
        for i in context_items
    )
    try:
        answer = llm.chat(
            "你是企业实体情报助手。基于提供的实体卡片总结回答用户问题，不要编造卡片之外的信息。"
            "回答要结构化：归属公司/参加的及关联的项目/项目进度/客户需求/沟通要点。",
            f"{context}\n\n【问题】{question}",
        )
    except RuntimeError as e:
        raise HTTPException(status_code=502, detail=f"AI 调用失败: {e}")

    return {
        "answer": answer,
        "sources": [{"id": i["id"], "type": i["type"], "name": i["name"]} for i in context_items[:6]],
    }
